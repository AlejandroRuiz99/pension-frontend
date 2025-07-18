"""
Módulo para generar archivos Excel con múltiples pestañas
Contiene toda la lógica de creación del archivo Excel con bases de cotización
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def generate_excel_from_process_result(result_data):
    """
    Genera un archivo Excel con múltiples pestañas a partir del resultado del procesamiento
    
    Args:
        result_data (dict): Datos del procesamiento con bases_procesadas y otros campos
        
    Returns:
        bytes: Contenido del archivo Excel en formato binario
    """
    try:
        # Crear workbook
        wb = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Eliminar hoja por defecto
        wb.remove(wb.active)
        
        # Filtrar bases por tipo
        bases_revalorizadas = [
            base for base in result_data.get("bases_procesadas", [])
            if base.get("periodo") == "revalorizado"
        ]
        
        bases_no_revalorizadas = [
            base for base in result_data.get("bases_procesadas", [])
            if base.get("periodo") == "no_revalorizado"
        ]
        
        # Crear las tres pestañas
        _create_bases_revalorizadas_sheet(wb, bases_revalorizadas, header_font, header_fill, border)
        _create_bases_no_revalorizadas_sheet(wb, bases_no_revalorizadas, header_font, header_fill, border)
        _create_resumen_sheet(wb, result_data, bases_revalorizadas, bases_no_revalorizadas, header_font, header_fill, border)
        
        # Crear buffer en memoria
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
        
    except Exception as e:
        print(f"Error generando Excel: {str(e)}")
        return None


def _create_bases_revalorizadas_sheet(wb, bases_revalorizadas, header_font, header_fill, border):
    """Crear pestaña de Bases Revalorizadas"""
    ws_revalorizadas = wb.create_sheet("Bases Revalorizadas")
    
    # Headers
    headers_revalorizadas = ["Mes/Año", "Base €", "Base Original €", "Índice", "Empresa", "Régimen", "Días Cotizados"]
    
    # Escribir headers
    for col, header in enumerate(headers_revalorizadas, 1):
        cell = ws_revalorizadas.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    # Escribir datos
    for row, base in enumerate(bases_revalorizadas, 2):
        ws_revalorizadas.cell(row=row, column=1, value=base.get("mes_anyo", ""))
        ws_revalorizadas.cell(row=row, column=2, value=base.get("base", 0))
        ws_revalorizadas.cell(row=row, column=3, value=base.get("base_original", 0))
        ws_revalorizadas.cell(row=row, column=4, value=base.get("indice_revalorizacion", 1))
        ws_revalorizadas.cell(row=row, column=5, value=base.get("empresa", ""))
        ws_revalorizadas.cell(row=row, column=6, value=base.get("regimen", ""))
        ws_revalorizadas.cell(row=row, column=7, value=30)  # Días cotizados
        
        # Aplicar bordes
        for col in range(1, 8):
            ws_revalorizadas.cell(row=row, column=col).border = border
    
    # Fórmulas para sumas totales
    if bases_revalorizadas:
        suma_row = len(bases_revalorizadas) + 3
        
        # Suma total bases
        ws_revalorizadas.cell(row=suma_row, column=1, value="SUMA TOTAL REVALORIZADAS:")
        ws_revalorizadas.cell(row=suma_row, column=1).font = Font(bold=True)
        ws_revalorizadas.cell(row=suma_row, column=2, value=f"=SUM(B2:B{len(bases_revalorizadas)+1})")
        ws_revalorizadas.cell(row=suma_row, column=2).font = Font(bold=True)
        ws_revalorizadas.cell(row=suma_row, column=2).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # Suma total días
        ws_revalorizadas.cell(row=suma_row, column=6, value="TOTAL DÍAS:")
        ws_revalorizadas.cell(row=suma_row, column=6).font = Font(bold=True)
        ws_revalorizadas.cell(row=suma_row, column=7, value=f"=SUM(G2:G{len(bases_revalorizadas)+1})")
        ws_revalorizadas.cell(row=suma_row, column=7).font = Font(bold=True)
        ws_revalorizadas.cell(row=suma_row, column=7).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # Suma días últimos 13 años (156 meses más recientes)
        ws_revalorizadas.cell(row=suma_row+1, column=6, value="ÚLTIMOS 13 AÑOS:")
        ws_revalorizadas.cell(row=suma_row+1, column=6).font = Font(bold=True)
        
        total_bases = len(bases_revalorizadas)
        if total_bases >= 156:
            # Sumar solo los 156 más recientes
            ws_revalorizadas.cell(row=suma_row+1, column=7, value=f"=SUM(G2:G157)")
        else:
            # Si hay menos de 156 bases, sumar todas
            ws_revalorizadas.cell(row=suma_row+1, column=7, value=f"=SUM(G2:G{total_bases+1})")
        
        ws_revalorizadas.cell(row=suma_row+1, column=7).font = Font(bold=True)
        ws_revalorizadas.cell(row=suma_row+1, column=7).fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    
    # Ajustar ancho de columnas
    for column in ws_revalorizadas.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_revalorizadas.column_dimensions[column_letter].width = adjusted_width


def _create_bases_no_revalorizadas_sheet(wb, bases_no_revalorizadas, header_font, header_fill, border):
    """Crear pestaña de Bases No Revalorizadas"""
    ws_no_revalorizadas = wb.create_sheet("Bases No Revalorizadas")
    
    # Headers
    headers_no_revalorizadas = ["Mes/Año", "Base €", "Empresa", "Régimen", "Días Cotizados"]
    
    # Escribir headers
    for col, header in enumerate(headers_no_revalorizadas, 1):
        cell = ws_no_revalorizadas.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    # Escribir datos
    for row, base in enumerate(bases_no_revalorizadas, 2):
        ws_no_revalorizadas.cell(row=row, column=1, value=base.get("mes_anyo", ""))
        ws_no_revalorizadas.cell(row=row, column=2, value=base.get("base", 0))
        ws_no_revalorizadas.cell(row=row, column=3, value=base.get("empresa", ""))
        ws_no_revalorizadas.cell(row=row, column=4, value=base.get("regimen", ""))
        ws_no_revalorizadas.cell(row=row, column=5, value=30)  # Días cotizados
        
        # Aplicar bordes
        for col in range(1, 6):
            ws_no_revalorizadas.cell(row=row, column=col).border = border
    
    # Fórmulas para sumas totales
    if bases_no_revalorizadas:
        suma_row = len(bases_no_revalorizadas) + 3
        
        # Suma total bases
        ws_no_revalorizadas.cell(row=suma_row, column=1, value="SUMA TOTAL NO REVALORIZADAS:")
        ws_no_revalorizadas.cell(row=suma_row, column=1).font = Font(bold=True)
        ws_no_revalorizadas.cell(row=suma_row, column=2, value=f"=SUM(B2:B{len(bases_no_revalorizadas)+1})")
        ws_no_revalorizadas.cell(row=suma_row, column=2).font = Font(bold=True)
        ws_no_revalorizadas.cell(row=suma_row, column=2).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Suma total días
        ws_no_revalorizadas.cell(row=suma_row, column=4, value="TOTAL DÍAS:")
        ws_no_revalorizadas.cell(row=suma_row, column=4).font = Font(bold=True)
        ws_no_revalorizadas.cell(row=suma_row, column=5, value=f"=SUM(E2:E{len(bases_no_revalorizadas)+1})")
        ws_no_revalorizadas.cell(row=suma_row, column=5).font = Font(bold=True)
        ws_no_revalorizadas.cell(row=suma_row, column=5).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Ajustar ancho de columnas
    for column in ws_no_revalorizadas.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_no_revalorizadas.column_dimensions[column_letter].width = adjusted_width


def _create_resumen_sheet(wb, result_data, bases_revalorizadas, bases_no_revalorizadas, header_font, header_fill, border):
    """Crear pestaña de Resumen y Cálculos"""
    ws_resumen = wb.create_sheet("Resumen y Cálculos")
    
    # Información general
    ws_resumen.cell(row=1, column=1, value="RESUMEN DEL CÁLCULO DE BASE REGULADORA")
    ws_resumen.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_resumen.merge_cells('A1:D1')
    
    row_counter = 3
    
    # Datos del expediente
    ws_resumen.cell(row=row_counter, column=1, value="DATOS DEL EXPEDIENTE")
    ws_resumen.cell(row=row_counter, column=1).font = Font(bold=True)
    ws_resumen.cell(row=row_counter, column=1).fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    row_counter += 1
    
    ws_resumen.cell(row=row_counter, column=1, value="Fecha de Jubilación:")
    ws_resumen.cell(row=row_counter, column=2, value=result_data.get("fecha_jubilacion", ""))
    row_counter += 1
    
    ws_resumen.cell(row=row_counter, column=1, value="Régimen de Acceso:")
    ws_resumen.cell(row=row_counter, column=2, value=result_data.get("regimen_acceso", ""))
    row_counter += 1
    
    ws_resumen.cell(row=row_counter, column=1, value="Sexo:")
    ws_resumen.cell(row=row_counter, column=2, value=result_data.get("sexo", ""))
    row_counter += 2
    
    # Estadísticas
    estadisticas = result_data.get("estadisticas", {})
    
    ws_resumen.cell(row=row_counter, column=1, value="ESTADÍSTICAS DE BASES")
    ws_resumen.cell(row=row_counter, column=1).font = Font(bold=True)
    ws_resumen.cell(row=row_counter, column=1).fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
    row_counter += 1
    
    ws_resumen.cell(row=row_counter, column=1, value="Total de Bases:")
    ws_resumen.cell(row=row_counter, column=2, value=estadisticas.get("total_bases", 0))
    row_counter += 1
    
    ws_resumen.cell(row=row_counter, column=1, value="Bases Revalorizadas:")
    ws_resumen.cell(row=row_counter, column=2, value=estadisticas.get("bases_revalorizadas", 0))
    row_counter += 1
    
    ws_resumen.cell(row=row_counter, column=1, value="Bases No Revalorizadas:")
    ws_resumen.cell(row=row_counter, column=2, value=estadisticas.get("bases_no_revalorizadas", 0))
    row_counter += 2
    
    # Cálculos de sumas dinámicos
    ws_resumen.cell(row=row_counter, column=1, value="CÁLCULOS DE SUMAS")
    ws_resumen.cell(row=row_counter, column=1).font = Font(bold=True)
    ws_resumen.cell(row=row_counter, column=1).fill = PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid")
    row_counter += 1
    
    # Referencias dinámicas a las sumas de las otras hojas
    ws_resumen.cell(row=row_counter, column=1, value="Suma Bases Revalorizadas:")
    if bases_revalorizadas:
        suma_rev_row_destino = len(bases_revalorizadas) + 3
        ws_resumen.cell(row=row_counter, column=2, value=f"='Bases Revalorizadas'!B{suma_rev_row_destino}")
    else:
        ws_resumen.cell(row=row_counter, column=2, value=0)
    suma_rev_row = row_counter
    row_counter += 1
    
    ws_resumen.cell(row=row_counter, column=1, value="Suma Bases No Revalorizadas:")
    if bases_no_revalorizadas:
        suma_no_rev_row_destino = len(bases_no_revalorizadas) + 3
        ws_resumen.cell(row=row_counter, column=2, value=f"='Bases No Revalorizadas'!B{suma_no_rev_row_destino}")
    else:
        ws_resumen.cell(row=row_counter, column=2, value=0)
    suma_no_rev_row = row_counter
    row_counter += 1
    
    ws_resumen.cell(row=row_counter, column=1, value="SUMA TOTAL:")
    ws_resumen.cell(row=row_counter, column=2, value=f"=B{suma_rev_row}+B{suma_no_rev_row}")
    ws_resumen.cell(row=row_counter, column=1).font = Font(bold=True)
    ws_resumen.cell(row=row_counter, column=2).font = Font(bold=True)
    ws_resumen.cell(row=row_counter, column=2).fill = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid")
    suma_total_row = row_counter
    row_counter += 2
    
    # Días cotizados últimos 15 años
    ws_resumen.cell(row=row_counter, column=1, value="DÍAS COTIZADOS ÚLTIMOS 15 AÑOS:")
    
    # Nueva lógica: Todas las no revalorizadas + últimos 13 años de revalorizadas
    if bases_revalorizadas and bases_no_revalorizadas:
        suma_no_rev_dias_row = len(bases_no_revalorizadas) + 3
        suma_13_anos_row = len(bases_revalorizadas) + 4
        formula_15_anos = f"='Bases No Revalorizadas'!E{suma_no_rev_dias_row}+'Bases Revalorizadas'!G{suma_13_anos_row}"
        ws_resumen.cell(row=row_counter, column=2, value=formula_15_anos)
        ws_resumen.cell(row=row_counter, column=3, value="(2 años no rev + 13 años rev)")
    elif bases_no_revalorizadas:
        suma_no_rev_dias_row = len(bases_no_revalorizadas) + 3
        ws_resumen.cell(row=row_counter, column=2, value=f"='Bases No Revalorizadas'!E{suma_no_rev_dias_row}")
        ws_resumen.cell(row=row_counter, column=3, value="(solo no revalorizadas)")
    elif bases_revalorizadas:
        suma_13_anos_row = len(bases_revalorizadas) + 4
        ws_resumen.cell(row=row_counter, column=2, value=f"='Bases Revalorizadas'!G{suma_13_anos_row}")
        ws_resumen.cell(row=row_counter, column=3, value="(solo últimos 13 años rev)")
    else:
        ws_resumen.cell(row=row_counter, column=2, value=0)
        ws_resumen.cell(row=row_counter, column=3, value="(sin datos)")
    
    ws_resumen.cell(row=row_counter, column=1).font = Font(bold=True)
    ws_resumen.cell(row=row_counter, column=2).font = Font(bold=True)
    ws_resumen.cell(row=row_counter, column=2).fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
    row_counter += 2
    
    # Parámetros de cómputo
    parametros = result_data.get("parametros_computo", {})
    
    ws_resumen.cell(row=row_counter, column=1, value="PARÁMETROS DE CÓMPUTO")
    ws_resumen.cell(row=row_counter, column=1).font = Font(bold=True)
    ws_resumen.cell(row=row_counter, column=1).fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    row_counter += 1
    
    ws_resumen.cell(row=row_counter, column=1, value="Bases Incluidas:")
    ws_resumen.cell(row=row_counter, column=2, value=parametros.get("bases_incluidas", 0))
    row_counter += 1
    
    ws_resumen.cell(row=row_counter, column=1, value="Período (meses):")
    ws_resumen.cell(row=row_counter, column=2, value=parametros.get("periodo_meses", 0))
    row_counter += 1
    
    ws_resumen.cell(row=row_counter, column=1, value="Divisor Base Reguladora:")
    ws_resumen.cell(row=row_counter, column=2, value=parametros.get("divisor_base_reguladora", 0))
    row_counter += 2
    
    # Base reguladora final
    ws_resumen.cell(row=row_counter, column=1, value="BASE REGULADORA:")
    divisor_cell = f"B{row_counter-2}"
    ws_resumen.cell(row=row_counter, column=2, value=f"=B{suma_total_row}/{divisor_cell}")
    
    # Estilo especial para la base reguladora
    ws_resumen.cell(row=row_counter, column=1).font = Font(bold=True, size=12, color="FFFFFF")
    ws_resumen.cell(row=row_counter, column=2).font = Font(bold=True, size=12, color="FFFFFF")
    ws_resumen.cell(row=row_counter, column=1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws_resumen.cell(row=row_counter, column=2).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Ajustar ancho de columnas del resumen
    ws_resumen.column_dimensions['A'].width = 30
    ws_resumen.column_dimensions['B'].width = 20
    ws_resumen.column_dimensions['C'].width = 25 