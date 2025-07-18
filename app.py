"""
🏢 Frontend para API de Bases de Cotización
Aplicación Streamlit moderna y funcional
"""

import streamlit as st
import requests
import json
import pandas as pd
from datetime import datetime
import time
import io
import sys
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuración de la página
st.set_page_config(
    page_title="Simulador de Pensiones",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# URL de la API
API_BASE_URL = "https://pension-bases-api-e707c1384c99.herokuapp.com"

# CSS personalizado para un diseño moderno
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        color: white;
        text-align: center;
    }
    
    .feature-card {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        margin: 1rem 0;
        border-left: 4px solid #667eea;
    }
    
    .success-box {
        background: #d4edda;
        border: 1px solid #c3e6cb;
        color: #155724;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
    
    .error-box {
        background: #f8d7da;
        border: 1px solid #f5c6cb;
        color: #721c24;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
    
    .stButton > button {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 5px;
        padding: 0.5rem 1rem;
        transition: all 0.3s;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 15px rgba(0,0,0,0.2);
    }
    
    .download-btn {
        background: #28a745 !important;
    }
    
    .excel-btn {
        background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important;
    }
    
    .debug-btn {
        background: linear-gradient(90deg, #6c757d 0%, #495057 100%) !important;
        font-size: 0.9em !important;
    }
</style>
""", unsafe_allow_html=True)

# Funciones para interactuar con la API
def check_api_health():
    """Verificar el estado de la API"""
    try:
        response = requests.get(f"{API_BASE_URL}/health", timeout=10)
        return response.status_code == 200, response.json() if response.status_code == 200 else None
    except Exception as e:
        return False, str(e)

def extract_bases(file):
    """Extraer bases de cotización de un PDF"""
    try:
        files = {"file": file}
        response = requests.post(f"{API_BASE_URL}/api/extract", files=files, timeout=60)
        return response.status_code == 200, response.json()
    except Exception as e:
        return False, {"error": str(e)}

def process_complete(file, fecha_jubilacion, regimen_acceso, sexo):
    """Procesar PDF completo (extraer + simular)"""
    try:
        files = {"file": file}
        data = {
            "fecha_jubilacion": fecha_jubilacion,
            "regimen_acceso": regimen_acceso,
            "sexo": sexo
        }
        response = requests.post(f"{API_BASE_URL}/api/process", files=files, data=data, timeout=120)
        return response.status_code == 200, response.json()
    except Exception as e:
        return False, {"error": str(e)}

def get_configuration():
    """Obtener configuración de la API"""
    try:
        # Obtener todas las configuraciones
        configs = {}
        
        # Parámetros de cómputo
        response = requests.get(f"{API_BASE_URL}/api/config/parametros", timeout=10)
        if response.status_code == 200:
            configs["parametros"] = response.json()
        
        # Índices de revalorización
        response = requests.get(f"{API_BASE_URL}/api/config/indices", timeout=10)
        if response.status_code == 200:
            configs["indices"] = response.json()
        
        # Topes de cotización
        response = requests.get(f"{API_BASE_URL}/api/config/topes", timeout=10)
        if response.status_code == 200:
            configs["topes"] = response.json()
        
        return True, configs
    except Exception as e:
        return False, {"error": str(e)}

def generate_excel_from_process_result(result_data):
    """
    Genera un archivo Excel con múltiples pestañas a partir del resultado del procesamiento
    """
    try:
        # Crear workbook
        wb = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Eliminar hoja por defecto
        wb.remove(wb.active)
        
        # =================== PESTAÑA 1: BASES REVALORIZADAS ===================
        ws_revalorizadas = wb.create_sheet("Bases Revalorizadas")
        
        # Filtrar bases revalorizadas
        bases_revalorizadas = [
            base for base in result_data.get("bases_procesadas", [])
            if base.get("periodo") == "revalorizado"
        ]
        
        # Headers para bases revalorizadas
        headers_revalorizadas = ["Mes/Año", "Base €", "Base Original €", "Índice", "Empresa", "Régimen", "Días Cotizados"]
        
        # Escribir headers
        for col, header in enumerate(headers_revalorizadas, 1):
            cell = ws_revalorizadas.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # Escribir datos de bases revalorizadas
        for row, base in enumerate(bases_revalorizadas, 2):
            ws_revalorizadas.cell(row=row, column=1, value=base.get("mes_anyo", ""))
            ws_revalorizadas.cell(row=row, column=2, value=base.get("base", 0))
            ws_revalorizadas.cell(row=row, column=3, value=base.get("base_original", 0))
            ws_revalorizadas.cell(row=row, column=4, value=base.get("indice_revalorizacion", 1))
            ws_revalorizadas.cell(row=row, column=5, value=base.get("empresa", ""))
            ws_revalorizadas.cell(row=row, column=6, value=base.get("regimen", ""))
            ws_revalorizadas.cell(row=row, column=7, value=30)  # Días cotizados
            
            # Aplicar bordes
            for col in range(1, 8):
                ws_revalorizadas.cell(row=row, column=col).border = border
        
        # Fórmulas para sumas totales
        if bases_revalorizadas:
            suma_row = len(bases_revalorizadas) + 3
            ws_revalorizadas.cell(row=suma_row, column=1, value="SUMA TOTAL REVALORIZADAS:")
            ws_revalorizadas.cell(row=suma_row, column=1).font = Font(bold=True)
            ws_revalorizadas.cell(row=suma_row, column=2, value=f"=SUM(B2:B{len(bases_revalorizadas)+1})")
            ws_revalorizadas.cell(row=suma_row, column=2).font = Font(bold=True)
            ws_revalorizadas.cell(row=suma_row, column=2).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            
            # Suma de días cotizados totales
            ws_revalorizadas.cell(row=suma_row, column=6, value="TOTAL DÍAS:")
            ws_revalorizadas.cell(row=suma_row, column=6).font = Font(bold=True)
            ws_revalorizadas.cell(row=suma_row, column=7, value=f"=SUM(G2:G{len(bases_revalorizadas)+1})")
            ws_revalorizadas.cell(row=suma_row, column=7).font = Font(bold=True)
            ws_revalorizadas.cell(row=suma_row, column=7).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            
            # Suma de días cotizados ÚLTIMOS 13 AÑOS (156 meses más recientes)
            ws_revalorizadas.cell(row=suma_row+1, column=6, value="ÚLTIMOS 13 AÑOS:")
            ws_revalorizadas.cell(row=suma_row+1, column=6).font = Font(bold=True)
            
            # Calcular los últimos 156 meses (13 años × 12 meses)
            total_bases = len(bases_revalorizadas)
            if total_bases >= 156:
                # Sumar solo los 156 más recientes (desde fila 2 hasta fila 157)
                ws_revalorizadas.cell(row=suma_row+1, column=7, value=f"=SUM(G2:G157)")
            else:
                # Si hay menos de 156 bases, sumar todas
                ws_revalorizadas.cell(row=suma_row+1, column=7, value=f"=SUM(G2:G{total_bases+1})")
            
            ws_revalorizadas.cell(row=suma_row+1, column=7).font = Font(bold=True)
            ws_revalorizadas.cell(row=suma_row+1, column=7).fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
        
        # Ajustar ancho de columnas
        for column in ws_revalorizadas.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_revalorizadas.column_dimensions[column_letter].width = adjusted_width
        
        # =================== PESTAÑA 2: BASES NO REVALORIZADAS ===================
        ws_no_revalorizadas = wb.create_sheet("Bases No Revalorizadas")
        
        # Filtrar bases no revalorizadas
        bases_no_revalorizadas = [
            base for base in result_data.get("bases_procesadas", [])
            if base.get("periodo") == "no_revalorizado"
        ]
        
        # Headers para bases no revalorizadas
        headers_no_revalorizadas = ["Mes/Año", "Base €", "Empresa", "Régimen", "Días Cotizados"]
        
        # Escribir headers
        for col, header in enumerate(headers_no_revalorizadas, 1):
            cell = ws_no_revalorizadas.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # Escribir datos de bases no revalorizadas
        for row, base in enumerate(bases_no_revalorizadas, 2):
            ws_no_revalorizadas.cell(row=row, column=1, value=base.get("mes_anyo", ""))
            ws_no_revalorizadas.cell(row=row, column=2, value=base.get("base", 0))
            ws_no_revalorizadas.cell(row=row, column=3, value=base.get("empresa", ""))
            ws_no_revalorizadas.cell(row=row, column=4, value=base.get("regimen", ""))
            ws_no_revalorizadas.cell(row=row, column=5, value=30)  # Días cotizados
            
            # Aplicar bordes
            for col in range(1, 6):
                ws_no_revalorizadas.cell(row=row, column=col).border = border
        
        # Fórmulas para sumas totales
        if bases_no_revalorizadas:
            suma_row = len(bases_no_revalorizadas) + 3
            ws_no_revalorizadas.cell(row=suma_row, column=1, value="SUMA TOTAL NO REVALORIZADAS:")
            ws_no_revalorizadas.cell(row=suma_row, column=1).font = Font(bold=True)
            ws_no_revalorizadas.cell(row=suma_row, column=2, value=f"=SUM(B2:B{len(bases_no_revalorizadas)+1})")
            ws_no_revalorizadas.cell(row=suma_row, column=2).font = Font(bold=True)
            ws_no_revalorizadas.cell(row=suma_row, column=2).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            
            # Suma de días cotizados
            ws_no_revalorizadas.cell(row=suma_row, column=4, value="TOTAL DÍAS:")
            ws_no_revalorizadas.cell(row=suma_row, column=4).font = Font(bold=True)
            ws_no_revalorizadas.cell(row=suma_row, column=5, value=f"=SUM(E2:E{len(bases_no_revalorizadas)+1})")
            ws_no_revalorizadas.cell(row=suma_row, column=5).font = Font(bold=True)
            ws_no_revalorizadas.cell(row=suma_row, column=5).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Ajustar ancho de columnas
        for column in ws_no_revalorizadas.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_no_revalorizadas.column_dimensions[column_letter].width = adjusted_width
        
        # =================== PESTAÑA 3: RESUMEN Y CÁLCULOS ===================
        ws_resumen = wb.create_sheet("Resumen y Cálculos")
        
        # Información general
        ws_resumen.cell(row=1, column=1, value="RESUMEN DEL CÁLCULO DE BASE REGULADORA")
        ws_resumen.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws_resumen.merge_cells('A1:D1')
        
        # Datos principales
        row_counter = 3
        
        # Información del expediente
        ws_resumen.cell(row=row_counter, column=1, value="DATOS DEL EXPEDIENTE")
        ws_resumen.cell(row=row_counter, column=1).font = Font(bold=True)
        ws_resumen.cell(row=row_counter, column=1).fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        row_counter += 1
        
        ws_resumen.cell(row=row_counter, column=1, value="Fecha de Jubilación:")
        ws_resumen.cell(row=row_counter, column=2, value=result_data.get("fecha_jubilacion", ""))
        row_counter += 1
        
        ws_resumen.cell(row=row_counter, column=1, value="Régimen de Acceso:")
        ws_resumen.cell(row=row_counter, column=2, value=result_data.get("regimen_acceso", ""))
        row_counter += 1
        
        ws_resumen.cell(row=row_counter, column=1, value="Sexo:")
        ws_resumen.cell(row=row_counter, column=2, value=result_data.get("sexo", ""))
        row_counter += 2
        
        # Estadísticas
        estadisticas = result_data.get("estadisticas", {})
        
        ws_resumen.cell(row=row_counter, column=1, value="ESTADÍSTICAS DE BASES")
        ws_resumen.cell(row=row_counter, column=1).font = Font(bold=True)
        ws_resumen.cell(row=row_counter, column=1).fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
        row_counter += 1
        
        ws_resumen.cell(row=row_counter, column=1, value="Total de Bases:")
        ws_resumen.cell(row=row_counter, column=2, value=estadisticas.get("total_bases", 0))
        row_counter += 1
        
        ws_resumen.cell(row=row_counter, column=1, value="Bases Revalorizadas:")
        ws_resumen.cell(row=row_counter, column=2, value=estadisticas.get("bases_revalorizadas", 0))
        row_counter += 1
        
        ws_resumen.cell(row=row_counter, column=1, value="Bases No Revalorizadas:")
        ws_resumen.cell(row=row_counter, column=2, value=estadisticas.get("bases_no_revalorizadas", 0))
        row_counter += 2
        
        # Cálculos - usar valores calculados para evitar problemas de fórmulas
        ws_resumen.cell(row=row_counter, column=1, value="CÁLCULOS DE SUMAS")
        ws_resumen.cell(row=row_counter, column=1).font = Font(bold=True)
        ws_resumen.cell(row=row_counter, column=1).fill = PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid")
        row_counter += 1
        
        # Crear referencias dinámicas a las sumas de las otras hojas
        ws_resumen.cell(row=row_counter, column=1, value="Suma Bases Revalorizadas:")
        if bases_revalorizadas:
            # Referencias dinámicas a la suma total de la hoja "Bases Revalorizadas"
            suma_rev_row_destino = len(bases_revalorizadas) + 3
            ws_resumen.cell(row=row_counter, column=2, value=f"='Bases Revalorizadas'!B{suma_rev_row_destino}")
        else:
            ws_resumen.cell(row=row_counter, column=2, value=0)
        suma_rev_row = row_counter  # Guardar referencia para fórmula
        row_counter += 1
        
        ws_resumen.cell(row=row_counter, column=1, value="Suma Bases No Revalorizadas:")
        if bases_no_revalorizadas:
            # Referencias dinámicas a la suma total de la hoja "Bases No Revalorizadas"
            suma_no_rev_row_destino = len(bases_no_revalorizadas) + 3
            ws_resumen.cell(row=row_counter, column=2, value=f"='Bases No Revalorizadas'!B{suma_no_rev_row_destino}")
        else:
            ws_resumen.cell(row=row_counter, column=2, value=0)
        suma_no_rev_row = row_counter  # Guardar referencia para fórmula
        row_counter += 1
        
        ws_resumen.cell(row=row_counter, column=1, value="SUMA TOTAL:")
        ws_resumen.cell(row=row_counter, column=2, value=f"=B{suma_rev_row}+B{suma_no_rev_row}")
        ws_resumen.cell(row=row_counter, column=1).font = Font(bold=True)
        ws_resumen.cell(row=row_counter, column=2).font = Font(bold=True)
        ws_resumen.cell(row=row_counter, column=2).fill = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid")
        suma_total_row = row_counter  # Guardar referencia para la base reguladora
        row_counter += 2
        
        # DÍAS COTIZADOS ÚLTIMOS 15 AÑOS
        ws_resumen.cell(row=row_counter, column=1, value="DÍAS COTIZADOS ÚLTIMOS 15 AÑOS:")
        
        # Nueva lógica: Todas las no revalorizadas + últimos 13 años de revalorizadas
        if bases_revalorizadas and bases_no_revalorizadas:
            # Fórmula dinámica: suma de no revalorizadas + últimos 13 años de revalorizadas
            suma_no_rev_row = len(bases_no_revalorizadas) + 3
            suma_13_anos_row = len(bases_revalorizadas) + 4  # +4 porque está en la segunda fila de totales
            formula_15_anos = f"='Bases No Revalorizadas'!E{suma_no_rev_row}+'Bases Revalorizadas'!G{suma_13_anos_row}"
            ws_resumen.cell(row=row_counter, column=2, value=formula_15_anos)
            ws_resumen.cell(row=row_counter, column=3, value="(2 años no rev + 13 años rev)")
            
        elif bases_no_revalorizadas:
            # Solo no revalorizadas
            suma_no_rev_row = len(bases_no_revalorizadas) + 3
            ws_resumen.cell(row=row_counter, column=2, value=f"='Bases No Revalorizadas'!E{suma_no_rev_row}")
            ws_resumen.cell(row=row_counter, column=3, value="(solo no revalorizadas)")
            
        elif bases_revalorizadas:
            # Solo revalorizadas - tomar últimos 13 años
            suma_13_anos_row = len(bases_revalorizadas) + 4
            ws_resumen.cell(row=row_counter, column=2, value=f"='Bases Revalorizadas'!G{suma_13_anos_row}")
            ws_resumen.cell(row=row_counter, column=3, value="(solo últimos 13 años rev)")
            
        else:
            # Sin datos
            ws_resumen.cell(row=row_counter, column=2, value=0)
            ws_resumen.cell(row=row_counter, column=3, value="(sin datos)")
            
        ws_resumen.cell(row=row_counter, column=1).font = Font(bold=True)
        ws_resumen.cell(row=row_counter, column=2).font = Font(bold=True)
        ws_resumen.cell(row=row_counter, column=2).fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
        dias_15_anos_row = row_counter  # Guardar referencia si se necesita
        row_counter += 2
        
        # Parámetros de cómputo
        parametros = result_data.get("parametros_computo", {})
        
        ws_resumen.cell(row=row_counter, column=1, value="PARÁMETROS DE CÓMPUTO")
        ws_resumen.cell(row=row_counter, column=1).font = Font(bold=True)
        ws_resumen.cell(row=row_counter, column=1).fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        row_counter += 1
        
        ws_resumen.cell(row=row_counter, column=1, value="Bases Incluidas:")
        ws_resumen.cell(row=row_counter, column=2, value=parametros.get("bases_incluidas", 0))
        row_counter += 1
        
        ws_resumen.cell(row=row_counter, column=1, value="Período (meses):")
        ws_resumen.cell(row=row_counter, column=2, value=parametros.get("periodo_meses", 0))
        row_counter += 1
        
        ws_resumen.cell(row=row_counter, column=1, value="Divisor Base Reguladora:")
        ws_resumen.cell(row=row_counter, column=2, value=parametros.get("divisor_base_reguladora", 0))
        row_counter += 2
        
        # BASE REGULADORA FINAL
        ws_resumen.cell(row=row_counter, column=1, value="BASE REGULADORA:")
        divisor_cell = f"B{row_counter-2}"     # Referencia al divisor (dos filas arriba)
        ws_resumen.cell(row=row_counter, column=2, value=f"=B{suma_total_row}/{divisor_cell}")
        
        # Estilo especial para la base reguladora
        ws_resumen.cell(row=row_counter, column=1).font = Font(bold=True, size=12)
        ws_resumen.cell(row=row_counter, column=2).font = Font(bold=True, size=12)
        ws_resumen.cell(row=row_counter, column=1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        ws_resumen.cell(row=row_counter, column=2).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        ws_resumen.cell(row=row_counter, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        ws_resumen.cell(row=row_counter, column=2).font = Font(bold=True, size=12, color="FFFFFF")
        
        # Ajustar ancho de columnas del resumen
        ws_resumen.column_dimensions['A'].width = 30
        ws_resumen.column_dimensions['B'].width = 20
        ws_resumen.column_dimensions['C'].width = 25
        
        # Crear buffer en memoria
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
        
    except Exception as e:
        st.error(f"Error generando Excel: {str(e)}")
        return None

# Header principal
st.markdown("""
<div class="main-header">
    <h1>Simulador de Pensiones</h1>
    <p>Extracción y cálculo de bases reguladoras profesional</p>
</div>
""", unsafe_allow_html=True)

# Verificar estado de la API
with st.sidebar:
    st.header("🔗 Estado de la API")
    
    if st.button("🔄 Verificar Conexión"):
        with st.spinner("Verificando conexión..."):
            is_healthy, health_data = check_api_health()
        
        if is_healthy:
            st.success("✅ API conectada correctamente")
            if health_data:
                st.json(health_data.get("services", {}))
        else:
            st.error("❌ Error de conexión con la API")
            if health_data:
                st.error(health_data)

# Navegación principal
st.sidebar.markdown("---")
option = st.sidebar.selectbox(
    "🎛️ Selecciona una opción:",
    ["🏠 Inicio", "📄 Extraer Bases", "🚀 Procesar Completo", "⚙️ Configuración"]
)

# Página de inicio
if option == "🏠 Inicio":
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("""
        <div class="feature-card">
            <h3>📄 Extracción</h3>
            <p>Sube un archivo PDF y extrae automáticamente las bases de cotización.</p>
            <ul>
                <li>Identifica empresas y regímenes</li>
                <li>Valida datos automáticamente</li>
                <li>Descarga resultados en JSON</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="feature-card">
            <h3>🚀 Procesamiento Completo</h3>
            <p>Extrae bases y calcula la base reguladora en un solo paso.</p>
            <ul>
                <li>Simulación de períodos futuros</li>
                <li>Aplicación de revalorización</li>
                <li>Cálculo de base reguladora</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div class="feature-card">
            <h3>⚙️ Configuración</h3>
            <p>Consulta los parámetros de cálculo actualizados.</p>
            <ul>
                <li>Índices de revalorización</li>
                <li>Topes de cotización</li>
                <li>Parámetros de cómputo</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

# Página de extracción
elif option == "📄 Extraer Bases":
    st.header("📄 Extracción de Bases de Cotización")
    
    st.markdown("""
    <div class="feature-card">
        <p>Sube un archivo PDF con bases de cotización de la Seguridad Social y obtén los datos estructurados automáticamente.</p>
    </div>
    """, unsafe_allow_html=True)
    
    uploaded_file = st.file_uploader(
        "Selecciona un archivo PDF",
        type=['pdf'],
        help="Archivo PDF con bases de cotización de la Seguridad Social (máximo 10MB)"
    )
    
    if uploaded_file is not None:
        # Mostrar información del archivo
        st.info(f"📁 Archivo: {uploaded_file.name} ({uploaded_file.size / 1024:.1f} KB)")
        
        if st.button("🔍 Extraer Bases", key="extract"):
            if uploaded_file.size > 10 * 1024 * 1024:  # 10MB
                st.error("❌ El archivo es demasiado grande. Máximo 10MB.")
            else:
                with st.spinner("Extrayendo bases de cotización..."):
                    success, result = extract_bases(uploaded_file)
                
                if success:
                    st.markdown("""
                    <div class="success-box">
                        ✅ <strong>Extracción completada exitosamente</strong>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Mostrar estadísticas
                    if "total_bases" in result:
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Total Bases", result["total_bases"])
                        with col2:
                            st.metric("Empresas", result.get("metadata", {}).get("total_empresas", "N/A"))
                        with col3:
                            periodo = result.get("metadata", {}).get("periodo_bases", {})
                            if periodo:
                                st.metric("Período", f"{periodo.get('desde', '')} - {periodo.get('hasta', '')}")
                    
                    # Botón de descarga
                    json_data = json.dumps(result, indent=2, ensure_ascii=False)
                    st.download_button(
                        label="📥 Descargar Resultados (JSON)",
                        data=json_data,
                        file_name=f"bases_extraidas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                        mime="application/json",
                        key="download_extract"
                    )
                else:
                    st.markdown(f"""
                    <div class="error-box">
                        ❌ <strong>Error en la extracción:</strong><br>
                        {result.get('error', result.get('detail', 'Error desconocido'))}
                    </div>
                    """, unsafe_allow_html=True)

# Página de procesamiento completo
elif option == "🚀 Procesar Completo":
    st.header("🚀 Procesamiento Completo")
    
    st.markdown("""
    <div class="feature-card">
        <p>Proceso completo en un solo paso: extrae bases del PDF y calcula automáticamente la base reguladora.</p>
    </div>
    """, unsafe_allow_html=True)
    
    uploaded_file = st.file_uploader(
        "Selecciona un archivo PDF",
        type=['pdf'],
        help="Archivo PDF con bases de cotización de la Seguridad Social (máximo 10MB)",
        key="process_file"
    )
    
    col1, col2 = st.columns(2)
    
    with col1:
        fecha_jubilacion = st.text_input(
            "📅 Fecha de Jubilación",
            placeholder="MM/YYYY",
            help="Formato: MM/YYYY (ejemplo: 06/2025)"
        )
        
        regimen_acceso = st.selectbox(
            "⚙️ Régimen de Acceso",
            ["GENERAL", "AUTONOMO"],
            help="Régimen de cotización"
        )
    
    with col2:
        sexo = st.selectbox(
            "👤 Sexo del Cotizante",
            ["MASCULINO", "FEMENINO"],
            help="Necesario para el cálculo correcto de lagunas"
        )
    
    if uploaded_file is not None and fecha_jubilacion:
        # Validar formato de fecha
        import re
        if not re.match(r'^\d{2}/\d{4}$', fecha_jubilacion):
            st.error("❌ Formato de fecha incorrecto. Use MM/YYYY")
        else:
            st.info(f"📁 Archivo: {uploaded_file.name} ({uploaded_file.size / 1024:.1f} KB)")
            
            if st.button("🚀 Procesar Completo", key="process"):
                if uploaded_file.size > 10 * 1024 * 1024:  # 10MB
                    st.error("❌ El archivo es demasiado grande. Máximo 10MB.")
                else:
                    with st.spinner("Procesando archivo completo... Esto puede tomar unos minutos."):
                        success, result = process_complete(uploaded_file, fecha_jubilacion, regimen_acceso, sexo)
                    
                    if success:
                        st.markdown("""
                        <div class="success-box">
                            ✅ <strong>Procesamiento completado exitosamente</strong>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        # Mostrar resultados principales
                        if "estadisticas" in result:
                            stats = result["estadisticas"]
                            
                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.metric("Total Bases", stats.get("total_bases", "N/A"))
                            with col2:
                                st.metric("Base Reguladora", f"€{stats.get('base_reguladora', 0):.2f}")
                            with col3:
                                st.metric("Suma Total", f"€{stats.get('suma_total', 0):.2f}")
                            with col4:
                                st.metric("Cálculo Usado", result.get("calculo_elegido", "N/A"))
                        
                        # Información adicional
                        st.subheader("📊 Detalles del Procesamiento")
                        info_col1, info_col2 = st.columns(2)
                        
                        with info_col1:
                            st.write(f"**Fecha de Jubilación:** {result.get('fecha_jubilacion', 'N/A')}")
                            st.write(f"**Régimen:** {result.get('regimen_acceso', 'N/A')}")
                            st.write(f"**Sexo:** {result.get('sexo', 'N/A')}")
                        
                        with info_col2:
                            if "metadata_extraccion" in result:
                                metadata = result["metadata_extraccion"]
                                st.write(f"**Bases Extraídas:** {metadata.get('total_bases_extraidas', 'N/A')}")
                                st.write(f"**Empresas:** {metadata.get('total_empresas', 'N/A')}")
                                periodo = metadata.get('periodo_extraido', {})
                                if periodo:
                                    st.write(f"**Período:** {periodo.get('desde', '')} - {periodo.get('hasta', '')}")
                        
                        # Botones de descarga
                        st.subheader("📥 Descargar Resultados")
                        
                        col_download1, col_download2 = st.columns(2)
                        
                        with col_download1:
                            # Generar Excel
                            excel_data = generate_excel_from_process_result(result)
                            if excel_data:
                                st.download_button(
                                    label="📊 Descargar Excel Editable",
                                    data=excel_data,
                                    file_name=f"bases_cotizacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="download_excel",
                                    help="Excel con pestañas separadas y fórmulas dinámicas para edición"
                                )
                        
                        with col_download2:
                            # JSON para desarrollo
                            json_data = json.dumps(result, indent=2, ensure_ascii=False)
                            st.download_button(
                                label="🔨 Descargar JSON (Debug)",
                                data=json_data,
                                file_name=f"procesamiento_completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                                mime="application/json",
                                key="download_process",
                                help="Archivo JSON completo para desarrollo y depuración"
                            )
                    else:
                        st.markdown(f"""
                        <div class="error-box">
                            ❌ <strong>Error en el procesamiento:</strong><br>
                            {result.get('error', result.get('detail', 'Error desconocido'))}
                        </div>
                        """, unsafe_allow_html=True)

# Página de configuración
elif option == "⚙️ Configuración":
    st.header("⚙️ Configuración del Sistema")
    
    st.markdown("""
    <div class="feature-card">
        <p>Consulta los parámetros de configuración cargados en el sistema para cálculos y simulaciones.</p>
    </div>
    """, unsafe_allow_html=True)
    
    if st.button("🔄 Cargar Configuración"):
        with st.spinner("Cargando configuración..."):
            success, configs = get_configuration()
        
        if success:
            st.success("✅ Configuración cargada correctamente")
            
            # Parámetros de cómputo
            if "parametros" in configs:
                st.subheader("📊 Parámetros de Cómputo")
                params_data = configs["parametros"]["data"]
                
                # Verificar si ya tenemos los parámetros directamente o están anidados
                params = None
                if params_data:
                    if "parametros_computo_anual" in params_data:
                        params = params_data["parametros_computo_anual"]
                    else:
                        # Asumir que params_data ya contiene los parámetros directamente
                        params = params_data
                
                if params and isinstance(params, dict):
                    # Convertir a DataFrame para mejor visualización
                    df_params = []
                    for año, datos in params.items():
                        if isinstance(datos, dict):
                            df_params.append({
                                "Año": año,
                                "Bases Incluidas": datos.get("bases_incluidas", datos.get("numero_bases", "N/A")),
                                "Período (meses)": datos.get("periodo_meses", "N/A"),
                                "Divisor Base Reguladora": datos.get("divisor_base_reguladora", datos.get("divisor", "N/A"))
                            })
                    
                    if df_params:
                        # Ordenar por año descendente
                        df = pd.DataFrame(df_params)
                        df = df.sort_values('Año', ascending=False)
                        st.dataframe(df, use_container_width=True, hide_index=True)
                        
                        # Mostrar información adicional en columnas
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Total Años", len(df_params))
                        with col2:
                            último_año = max(params.keys())
                            st.metric("Último Año", último_año)
                        with col3:
                            if df_params:
                                primer_año = min(params.keys())
                                st.metric("Primer Año", primer_año)
                else:
                    st.warning("⚠️ No se pudieron procesar los parámetros de cómputo")
                    if params_data:
                        st.info("Los datos están disponibles pero en un formato no reconocido.")
            
            # Índices de revalorización
            if "indices" in configs:
                st.subheader("📈 Índices de Revalorización")
                indices_data = configs["indices"]["data"]
                
                # Verificar si ya tenemos los índices directamente o están anidados
                indices = None
                if indices_data:
                    if "indices_revalorizacion" in indices_data:
                        indices = indices_data["indices_revalorizacion"]
                    else:
                        # Asumir que indices_data ya contiene los índices directamente
                        indices = indices_data
                
                if indices and isinstance(indices, dict):
                    # Mostrar información resumida
                    total_indices = len(indices)
                    fechas = list(indices.keys())
                    if fechas:
                        desde = min(fechas)
                        hasta = max(fechas)
                        
                        # Métricas principales
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Total Índices", total_indices)
                        with col2:
                            st.metric("Desde", desde)
                        with col3:
                            st.metric("Hasta", hasta)
                    
                    # Mostrar todos los índices ordenados por fecha
                    def sort_key(fecha):
                        """Función para ordenar fechas MM/YYYY correctamente"""
                        try:
                            mes, año = fecha.split('/')
                            return (int(año), int(mes))
                        except:
                            return (0, 0)
                    
                    sorted_dates = sorted(indices.keys(), key=sort_key, reverse=True)
                    
                    df_indices = []
                    for fecha in sorted_dates:
                        indice = indices[fecha]
                        df_indices.append({
                            "Fecha (MM/YYYY)": fecha,
                            "Índice de Revalorización": f"{indice:.6f}"
                        })
                    
                    if df_indices:
                        st.write(f"**📊 Todos los índices de revalorización ({len(df_indices)} registros):**")
                        df = pd.DataFrame(df_indices)
                        
                        # Mostrar tabla con altura fija para scroll
                        st.dataframe(df, use_container_width=True, hide_index=True, height=400)
                        
                        # Estadísticas básicas
                        valores = list(indices.values())
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Índice Promedio", f"{sum(valores)/len(valores):.6f}")
                        with col2:
                            st.metric("Índice Máximo", f"{max(valores):.6f}")
                        with col3:
                            st.metric("Índice Mínimo", f"{min(valores):.6f}")
                        
                        # Información del último índice disponible
                        ultimo_fecha = sorted_dates[0]
                        ultimo_indice = indices[ultimo_fecha]
                        st.info(f"📅 **Último índice disponible**: {ultimo_fecha} - Valor: {ultimo_indice:.6f}")
                else:
                    st.warning("⚠️ No se pudieron procesar los índices de revalorización")
                    if indices_data:
                        st.info("Los datos están disponibles pero en un formato no reconocido.")
            
            # Topes de cotización
            if "topes" in configs:
                st.subheader("💰 Topes de Cotización")
                topes_data = configs["topes"]["data"]
                
                # Verificar si ya tenemos los topes directamente o están anidados
                topes = None
                if topes_data:
                    if "topes_cotizacion" in topes_data:
                        topes = topes_data["topes_cotizacion"]
                    else:
                        # Asumir que topes_data ya contiene los topes directamente
                        topes = topes_data
                
                if topes and isinstance(topes, dict):
                    df_topes = []
                    for año, datos in topes.items():
                        if isinstance(datos, dict):
                            base_min = datos.get('base_minima_mensual', 0)
                            base_max = datos.get('base_maxima_mensual', 0)
                            df_topes.append({
                                "Año": año,
                                "Base Mínima Mensual": f"€{base_min:.2f}",
                                "Base Máxima Mensual": f"€{base_max:.2f}",
                                "Diferencia": f"€{base_max - base_min:.2f}"
                            })
                    
                    if df_topes:
                        # Ordenar por año descendente
                        df = pd.DataFrame(df_topes)
                        df = df.sort_values('Año', ascending=False)
                        st.dataframe(df, use_container_width=True, hide_index=True)
                        
                        # Mostrar métricas en columnas
                        col1, col2, col3, col4 = st.columns(4)
                        
                        ultimo_año = max(topes.keys())
                        ultimo_datos = topes[ultimo_año]
                        
                        with col1:
                            st.metric("Total Años", len(df_topes))
                        with col2:
                            st.metric("Último Año", ultimo_año)
                        with col3:
                            st.metric("Mínima Actual", f"€{ultimo_datos.get('base_minima_mensual', 0):.2f}")
                        with col4:
                            st.metric("Máxima Actual", f"€{ultimo_datos.get('base_maxima_mensual', 0):.2f}")
                        
                        # Información destacada
                        st.success(f"🎯 **Bases de cotización para {ultimo_año}**: Desde €{ultimo_datos.get('base_minima_mensual', 0):.2f} hasta €{ultimo_datos.get('base_maxima_mensual', 0):.2f} mensuales")
                else:
                    st.warning("⚠️ No se pudieron procesar los topes de cotización")
                    if topes_data:
                        st.info("Los datos están disponibles pero en un formato no reconocido.")
            
            # Botón para descargar toda la configuración
            json_data = json.dumps(configs, indent=2, ensure_ascii=False)
            st.download_button(
                label="📥 Descargar Configuración Completa (JSON)",
                data=json_data,
                file_name=f"configuracion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json",
                key="download_config"
            )
        else:
            st.markdown(f"""
            <div class="error-box">
                ❌ <strong>Error cargando configuración:</strong><br>
                {configs.get('error', 'Error desconocido')}
            </div>
            """, unsafe_allow_html=True)

# Footer
st.markdown("---")
st.markdown(f"""
<div style="text-align: center; color: #666; padding: 1rem;">
    <p>🏢 API Bases de Cotización | Desarrollado con ❤️ usando Streamlit</p>
    <p>API actual: <strong>{API_BASE_URL}</strong></p>
    <p>💡 <em>Usa <code>streamlit run app.py --local</code> para conectar a API local (localhost:8000)</em></p>
</div>
""", unsafe_allow_html=True) 